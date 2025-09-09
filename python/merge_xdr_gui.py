import sys
import json
import re
from pathlib import Path
from shutil import copyfile
from typing import Any, List, Optional

import pandas as pd
from openpyxl import load_workbook

# ---- Qt ----
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import (
    QApplication, QWidget, QMainWindow, QPushButton, QLabel, QLineEdit, QTextEdit,
    QFileDialog, QMessageBox, QGridLayout
)

# ===================== 抽取阶段工具函数 =====================


def safe_render(v: Any) -> Any:
    if isinstance(v, dict) and "renderValue" in v:
        return v.get("renderValue")
    return v


def join_items(items: List[Any]) -> str:
    out = []
    for it in items:
        it = safe_render(it)
        if isinstance(it, (dict, list)):
            out.append(json.dumps(it, ensure_ascii=False))
        elif it is None:
            continue
        else:
            out.append(str(it))
    return ",".join(out)


def extract_http_header_value(raw_headers: str, key: str) -> str:
    if not raw_headers or not isinstance(raw_headers, str):
        return ""
    pattern = re.compile(
        rf"^{re.escape(key)}\s*:\s*(.*)$", re.IGNORECASE | re.MULTILINE)
    m = pattern.search(raw_headers)
    return m.group(1).strip() if m else ""


def sanitize_for_excel(val):
    if pd.isna(val):
        return val
    s = str(val)
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", " ", s)


def extract_dataframe_from_json(json_path: Path) -> pd.DataFrame:
    assert json_path.exists(), f"找不到文件：{json_path}"
    with json_path.open("r", encoding="utf-8") as f:
        root = json.load(f)

    records = root.get("data", {}).get("data", [])
    rows = []

    for rec in records:
        occur_ts = safe_render(rec.get("occurTime", ""))
        upload_ts = safe_render(rec.get("uploadTime", ""))
        insert_ts = safe_render(rec.get("insertTime", ""))
        uu_id = safe_render(rec.get("uuId", ""))
        platform_host_branch_id = safe_render(
            rec.get("platformHostBranchId", ""))

        # 目的端四列来源
        host_classify_id = safe_render(rec.get("hostClassifyId", ""))
        host_country = safe_render(rec.get("hostCountryName", ""))
        host_province = safe_render(rec.get("hostProvinceName", ""))
        host_asset_id = safe_render(rec.get("hostAssetId", ""))

        threat_class = safe_render(rec.get("threatClass", ""))
        alert_rule_id = safe_render(rec.get("alertRuleId", ""))
        rule_name = safe_render(rec.get("name", ""))
        rule_desc = safe_render(rec.get("description", ""))
        attack_result = safe_render(rec.get("attackResult", ""))
        white_status = safe_render(rec.get("whiteStatus", ""))
        deal_status = safe_render(rec.get("dealStatus", ""))
        severity_val = safe_render(rec.get("severity", ""))
        confidence_val = safe_render(rec.get("confidence", ""))
        recommendation_val = safe_render(rec.get("recommendation", ""))
        resp_status = safe_render(rec.get("respStatus", ""))
        request_head = safe_render(rec.get("requestHead", ""))
        response_head = safe_render(rec.get("responseHead", ""))
        request_body = safe_render(rec.get("requestBody", ""))
        response_body = safe_render(rec.get("responseBody", ""))
        req_cookie = safe_render(rec.get("reqCookie", ""))
        dns_queries = safe_render(rec.get("dnsQueries", ""))
        threat_sub_type = safe_render(rec.get("threatSubType", ""))

        # devices.data
        devs = rec.get("devices", {}).get("data", [])
        dev_vendors, dev_productTypes, dev_productVers = [], [], []
        dev_manage, dev_originTypes, dev_originVers, dev_deviceIds = [], [], [], []
        for d in devs:
            rv = safe_render(d)
            if not isinstance(rv, dict):
                continue
            dev_vendors.append(str(rv.get("vendor", "")))
            dev_productTypes.append(str(rv.get("productType", "")))
            dev_productVers.append(str(rv.get("productVer", "")))
            dev_manage.append(str(rv.get("manage", "")))
            dev_originTypes.append(str(rv.get("originProductType", "")))
            dev_originVers.append(str(rv.get("originProductVer", "")))
            dev_deviceIds.append(str(rv.get("deviceId", "")))

        vendor = ",".join([v for v in dev_vendors if v])
        productType = ",".join([v for v in dev_productTypes if v])
        productVer = ",".join([v for v in dev_productVers if v])
        manage = ",".join([v for v in dev_manage if v])
        originProductType = ",".join([v for v in dev_originTypes if v])
        originProductVer = ",".join([v for v in dev_originVers if v])
        deviceId = ",".join([v for v in dev_deviceIds if v])

        # srcIpInfos.data
        src_infos = rec.get("srcIpInfos", {}).get("data", [])
        src_ips, src_types, src_countries, src_provinces, src_asset_ids = [], [], [], [], []
        for s in src_infos:
            s_rv = safe_render(s)
            if isinstance(s_rv, dict):
                src_ips.append(str(s_rv.get("ip", "")))
                src_types.append(str(s_rv.get("classifyId", "")))
                src_countries.append(str(s_rv.get("countryName", "")))
                src_provinces.append(str(s_rv.get("provinceName", "")))
                src_asset_ids.append(str(s_rv.get("platformAssetId", "")))

        srcIp = ",".join([v for v in src_ips if v])
        srcType = ",".join([v for v in src_types if v])
        srcCountry = ",".join([v for v in src_countries if v])
        srcProvince = ",".join([v for v in src_provinces if v])
        srcAssetId = ",".join([v for v in src_asset_ids if v])

        # 其他列表字段
        srcPort = join_items(rec.get("srcPort", {}).get("data", []))
        dstIp = join_items(rec.get("dstIp", {}).get("data", []))

        # dstIpInfos.data -> 仅取 renderValue.platformAssetId，拼到 dstIpTag
        dst_infos = rec.get("dstIpInfos", {}).get("data", [])
        dst_platform_asset_ids: List[str] = []
        for d in dst_infos:
            d_rv = safe_render(d)
            if isinstance(d_rv, dict):
                pa = d_rv.get("platformAssetId", "")
                if pa:
                    dst_platform_asset_ids.append(str(pa))
        dstIpTag_val = ",".join(
            dst_platform_asset_ids) if dst_platform_asset_ids else ""

        # host* -> 目的端四列
        dstType_val = host_classify_id
        dstCountry_val = host_country
        dstProvince_val = host_province
        dstAssetId_val = host_asset_id

        dstPort = join_items(rec.get("dstPort", {}).get("data", []))
        engine = join_items(rec.get("engineName", {}).get("data", []))
        proofType_joined = join_items(rec.get("proofType", {}).get("data", []))
        technique = join_items(rec.get("attckTechnique", {}).get("data", []))
        url_joined = join_items(rec.get("url", {}).get("data", []))
        xff_joined = join_items(rec.get("xForwardedFor", {}).get("data", []))

        referer = extract_http_header_value(
            request_head, "Referer") if isinstance(request_head, str) else ""
        user_agent = extract_http_header_value(
            request_head, "User-Agent") if isinstance(request_head, str) else ""

        row = {
            "recordTimestamp": occur_ts,
            "uploadTimestamp": upload_ts,
            "insertTimestamp": insert_ts,
            "vendor": vendor,
            "productType": productType,
            "productVer": productVer,
            "manage": manage,
            "originProductType": originProductType,
            "originProductVer": originProductVer,
            "deviceId": deviceId,
            "uuId": uu_id,
            "srcIp": srcIp,
            "srcPort": srcPort,
            "srcIpTag": platform_host_branch_id,
            "srcType": srcType,
            "srcCountry": srcCountry,
            "srcProvince": srcProvince,
            "srcAssetId": srcAssetId,
            "dstIp": dstIp,
            "dstPort": dstPort,
            "dstIpTag": dstIpTag_val,
            "dstType": dstType_val,
            "dstCountry": dstCountry_val,
            "dstProvince": dstProvince_val,
            "dstAssetId": dstAssetId_val,
            "threatClass": threat_class,
            "reqRuleId": alert_rule_id,
            "ruleName": rule_name,
            "ruleDescription": rule_desc,
            "engine": engine,
            "l7Protocol": proofType_joined,
            "attackState": attack_result,
            "white": white_status,
            "action": deal_status,
            "technique": technique,
            "severity": severity_val,
            "riskLevel": severity_val,
            "confidence": confidence_val,
            "evidence": rule_desc,
            "suggestion": recommendation_val,
            "proofType": proofType_joined,
            "url": url_joined,
            "httpUri": url_joined,
            "referer": referer,
            "userAgent": user_agent,
            "xForwarded For": xff_joined,
            "respStatus": resp_status,
            "requestHead": request_head,
            "responseHead": response_head,
            "requestBody": request_body,
            "responseBody": response_body,
            "reqCookie": req_cookie,
            "xForwardedForTag": xff_joined,
            "dnsQueries": ",".join([str(x) for x in dns_queries]) if isinstance(dns_queries, list)
                          else (str(dns_queries) if dns_queries else ""),
            "virusFamily": rule_name,
            "virusName": rule_name,
            "virusType": threat_sub_type,
        }
        rows.append(row)

    columns_order = [
        "recordTimestamp", "uploadTimestamp", "insertTimestamp", "vendor", "productType", "productVer", "manage",
        "originProductType", "originProductVer", "deviceId", "uuId", "srcIp", "srcPort", "srcIpTag", "srcType",
        "srcCountry", "srcProvince", "srcAssetId", "dstIp", "dstPort", "dstIpTag", "dstType", "dstCountry", "dstProvince", "dstAssetId",
        "threatClass", "reqRuleId", "ruleName", "ruleDescription", "engine", "l7Protocol", "attackState", "white", "action", "technique",
        "severity", "riskLevel", "confidence", "evidence", "suggestion", "proofType", "url", "httpUri", "referer", "userAgent",
        "xForwarded For", "respStatus", "requestHead", "responseHead", "requestBody", "responseBody", "reqCookie", "xForwardedForTag",
        "dnsQueries", "virusFamily", "virusName", "virusType"
    ]
    df = pd.DataFrame(rows, columns=columns_order)
    return df

# ===================== 映射阶段工具函数 =====================


def norm(s: str) -> str:
    s = str(s)
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()


def to_int_like_if_possible(val):
    if pd.isna(val) or val == "":
        return ""
    if isinstance(val, int):
        return int(val)
    if isinstance(val, float):
        return int(val) if float(val).is_integer() else sanitize_for_excel(val)
    s = str(val).strip()
    if re.fullmatch(r"\d+\.0+", s):
        return int(s.split(".")[0])
    return sanitize_for_excel(s)


def fill_template_inplace(df: pd.DataFrame, template_path: Path, sheet_name: Optional[str] = None) -> Path:
    assert template_path.exists(), f"找不到模板文件：{template_path}"
    backup_path = template_path.with_name(template_path.stem + "_备份.xlsx")
    copyfile(template_path, backup_path)

    wb = load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    ncols = ws.max_column
    en_header = [(ws.cell(row=2, column=c).value or "")
                 for c in range(1, ncols + 1)]

    norm_map = {}
    for c in df.columns:
        key = norm(c)
        if key not in norm_map:
            norm_map[key] = c

    col_mapping = []
    for tgt in en_header:
        tgt_name = str(tgt)
        if tgt_name.strip() == "":
            col_mapping.append(None)
            continue
        src_name = norm_map.get(norm(tgt_name))
        col_mapping.append(src_name)

    out_rows = []
    for i in range(len(df)):
        row_vals = []
        for col_idx, src_name in enumerate(col_mapping):
            tgt_header = en_header[col_idx]
            tgt_norm = norm(tgt_header)
            if src_name is None:
                row_vals.append("")
                continue
            v = df.at[i, src_name]
            if tgt_norm == "respstatus":
                row_vals.append(to_int_like_if_possible(v))
            else:
                row_vals.append("" if pd.isna(v) else sanitize_for_excel(v))
        out_rows.append(row_vals)

    start_row = 3
    old_max_row = ws.max_row
    r = start_row
    for data_row in out_rows:
        for c in range(1, ncols + 1):
            val = data_row[c-1] if c-1 < len(data_row) else ""
            ws.cell(row=r, column=c, value=val)
        r += 1

    last_written_row = start_row + len(out_rows) - 1

    if old_max_row > last_written_row:
        for rr in range(last_written_row + 1, old_max_row + 1):
            for cc in range(1, ncols + 1):
                ws.cell(row=rr, column=cc, value="")

    wb.save(template_path)
    return backup_path

# ===================== GUI =====================


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("XDR 抽取与模板映射（PyQt5）")
        self.resize(800, 520)

        self.json_path: Optional[Path] = None
        self.tpl_path: Optional[Path] = None

        # --- 控件 ---
        self.btn_pick_json = QPushButton("读取（选择 JSON）")
        self.lbl_json = QLabel("未选择 JSON 文件")
        self.lbl_json.setStyleSheet("color: #555;")

        self.btn_pick_tpl = QPushButton("映射（选择模板Excel）")
        self.lbl_tpl = QLabel("未选择 模板文件（告警字段1.xlsx）")
        self.lbl_tpl.setStyleSheet("color: #555;")

        self.sheet_label = QLabel("工作表名（可选）：")
        self.sheet_edit = QLineEdit()
        self.sheet_edit.setPlaceholderText("不填则使用活动工作表")

        self.btn_run = QPushButton("转换（备份并写回）")
        self.btn_run.setEnabled(False)

        # ★ 新增：导出日志按钮
        self.btn_export_log = QPushButton("导出日志到文件")

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setPlaceholderText("日志输出...")

        # --- 布局 ---
        top = QWidget()
        self.setCentralWidget(top)
        grid = QGridLayout()
        top.setLayout(grid)

        grid.addWidget(self.btn_pick_json, 0, 0, 1, 1)
        grid.addWidget(self.lbl_json,      0, 1, 1, 3)

        grid.addWidget(self.btn_pick_tpl,  1, 0, 1, 1)
        grid.addWidget(self.lbl_tpl,       1, 1, 1, 3)

        grid.addWidget(self.sheet_label,   2, 0, 1, 1)
        grid.addWidget(self.sheet_edit,    2, 1, 1, 3)

        grid.addWidget(self.btn_run,       3, 0, 1, 2)
        grid.addWidget(self.btn_export_log, 3, 2, 1, 2)   # 放在同一行右侧

        grid.addWidget(self.log,           4, 0, 1, 4)

        # --- 事件 ---
        self.btn_pick_json.clicked.connect(self.pick_json)
        self.btn_pick_tpl.clicked.connect(self.pick_template)
        self.btn_run.clicked.connect(self.run_pipeline)
        self.btn_export_log.clicked.connect(self.export_log)  # ★ 绑定导出事件

    def append_log(self, text: str):
        self.log.append(text)
        self.log.moveCursor(self.log.textCursor().End)

    def pick_json(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 xdr告警日志.json", "", "JSON 文件 (*.json);;所有文件 (*)"
        )
        if path:
            self.json_path = Path(path)
            self.lbl_json.setText(str(self.json_path))
            self.append_log(f"[选择] JSON：{self.json_path}")
        self.refresh_run_button()

    def pick_template(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 模板（告警字段1.xlsx）", "", "Excel 文件 (*.xlsx);;所有文件 (*)"
        )
        if path:
            self.tpl_path = Path(path)
            self.lbl_tpl.setText(str(self.tpl_path))
            self.append_log(f"[选择] 模板：{self.tpl_path}")
        self.refresh_run_button()

    def refresh_run_button(self):
        self.btn_run.setEnabled(
            self.json_path is not None and self.tpl_path is not None)

    def run_pipeline(self):
        if not self.json_path or not self.tpl_path:
            QMessageBox.warning(self, "提示", "请先选择 JSON 与 模板文件。")
            return

        sheet_name = self.sheet_edit.text().strip() or None
        try:
            self.append_log("[开始] 抽取 JSON → DataFrame（不落盘）...")
            df = extract_dataframe_from_json(self.json_path)
            self.append_log(f"[完成] 抽取 {len(df)} 行。")

            self.append_log("[开始] 备份并写回模板（第3行起）...")
            backup = fill_template_inplace(
                df, self.tpl_path, sheet_name=sheet_name)
            self.append_log(f"[完成] 已写回模板：{self.tpl_path}")
            self.append_log(f"[备份] 已生成：{backup}")

            QMessageBox.information(self, "完成", "转换成功：已备份并写回模板。")
        except Exception as e:
            self.append_log(f"[错误] {e}")
            QMessageBox.critical(self, "错误", f"转换失败：{e}")

    # ★ 新增：导出日志功能
    def export_log(self):
        content = self.log.toPlainText()
        if not content.strip():
            QMessageBox.information(self, "提示", "当前日志为空。")
            return
        path, sel = QFileDialog.getSaveFileName(
            self, "导出日志", "", "文本文件 (*.txt);;日志文件 (*.log);;所有文件 (*)"
        )
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(content)
            self.append_log(f"[导出] 日志已保存到：{path}")
            QMessageBox.information(self, "完成", "日志已导出。")
        except Exception as e:
            self.append_log(f"[错误] 导出日志失败：{e}")
            QMessageBox.critical(self, "错误", f"导出日志失败：{e}")


def main():
    # 高分屏支持
    QApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)
    QApplication.setAttribute(QtCore.Qt.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
