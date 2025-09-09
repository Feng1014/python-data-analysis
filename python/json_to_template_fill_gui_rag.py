from __future__ import annotations

import base64
import json
import re
import os
import sys
import shutil
from typing import Any, Dict, List, Tuple, Optional, Iterable, Union
from copy import copy

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ------------------ 业务常量与解析逻辑 ------------------

PP_KEYS = [
    "tcp.payload",
    "http.request.method",
    "http.request.full_uri",
    "http.referer",
    "http.user_agent",
    "User-Agent",
    "http.content_type",
    "dns.qry.name",
    "dns.qry.type",
    "smb.cmd",
    "smb.flags",
    "smb.flags2",
    "smb.wct",
    "smb.pid",
    "smb.uid",
    "smb.mid",
    "ssh.protocol",
    "ftp.request.command",
    "ftp.request.arg",
    "ftp.response.code",
    "ftp.response.arg",
    "nbns.name",
    "nbns.id",
    "data-text-lines",
    "frame.protocols",
]

HEADERS = [
    "recordTimestamp", "uploadTimestamp", "insertTimestamp",
    "productType", "manage", "originProductType",
    "deviceId", "deviceIp", "uuId",
    "srcIp", "srcPort", "srcCountry", "srcProvince", "srcCity",
    "dstIp", "dstPort", "dstCountry", "dstProvince",
    "threatClass", "reqRuleId", "ruleName", "ruleDescription",
    "l4Protocol", "attackState", "action", "tactic", "technique",
    "riskLevel", "virusFamily", "virusName",
    "payload", "reqMethod", "url", "httpUri", "referer", "userAgent",
    "reqContentType", "dnsQueries", "dnsQTypes",
    "smbCommand", "smbFlags", "smbFlags2", "smbWordCount",
    "smbProcessId", "smbUserId", "smbMessageId",
    "sshProtocolVersion", "sshClientVersion", "ftpCommandMsg",
    "l7Protocol", "proofType",
    "severity", "confidence", "evidence", "packetData", "requestBody",
    "engine", "l3Protocol", "rawMsg",
]

_CANON_FAMILIES = {
    "emotet": "Emotet", "trickbot": "TrickBot", "qakbot": "Qakbot", "qbot": "Qakbot",
    "icedid": "IcedID", "ursnif": "Ursnif", "dridex": "Dridex", "gozi": "Gozi",
    "tinba": "Tinba", "zeus": "Zeus", "zbot": "Zeus", "ramnit": "Ramnit",
    "gamarue": "Gamarue", "conficker": "Conficker", "virut": "Virut", "sality": "Sality",
    "raccoon": "Raccoon", "vidar": "Vidar", "azorult": "Azorult", "redline": "RedLine",
    "formbook": "FormBook", "agenttesla": "AgentTesla", "lokibot": "LokiBot",
    "remcos": "Remcos", "nanocore": "NanoCore", "quasar": "Quasar",
    "asyncrat": "AsyncRAT", "njrat": "NjRAT", "warzone": "WarzoneRAT",
    "netwire": "NetWire", "poisonivy": "PoisonIvy", "plugx": "PlugX",
    "gh0st": "Gh0st", "shadowpad": "ShadowPad", "lockbit": "LockBit",
    "revil": "REvil", "sodinokibi": "REvil", "conti": "Conti", "ryuk": "Ryuk",
    "gandcrab": "GandCrab", "cerber": "Cerber", "blackcat": "BlackCat",
    "alphv": "BlackCat", "blackbasta": "BlackBasta", "clop": "Clop", "maze": "Maze",
    "hive": "Hive", "akira": "Akira", "royal": "Royal", "blackmatter": "BlackMatter",
    "ransomexx": "RansomEXX", "medusa": "Medusa", "xmrig": "XMRig",
    "kinsing": "Kinsing", "lemonduck": "LemonDuck", "mirai": "Mirai",
    "gafgyt": "Gafgyt", "mozi": "Mozi", "cobalt strike": "CobaltStrike",
    "cobaltstrike": "CobaltStrike", "metasploit": "Metasploit", "sliver": "Sliver"
}
_SUFFIX_KEYWORDS = [
    "checkin", "beacon", "beaconing", "c2", "cnc", "callback", "loader", "downloader",
    "dropper", "backdoor", "ransomware", "trojan", "worm", "miner", "coinminer", "botnet",
    "rat", "keylogger", "stealer", "infostealer", "spyware", "adware", "variant", "payload",
    "infection", "probe", "scanner"
]
_RE_FAMILY_STYLE = [
    re.compile(
        r'(?i)\b(?:win32|w32|msil|android|linux|osx|macos|js|vbs?|html|php|ps1?|powershell|elf|office|docm?|xlsm?|pdf)[/\.]([a-z0-9._-]{2,})'),
    re.compile(
        r'(?i)\b(?:trojan|worm|ransomware|backdoor|spyware|adware|rootkit|botnet|dropper|downloader)[/\.]([a-z0-9._-]{2,})'),
]


def _normalize_space(s: str) -> str:
    s = re.sub(r'[\u3000\t]+', ' ', s or '')
    s = re.sub(r'\s{2,}', ' ', s).strip()
    return s


def _canon(s: str) -> str:
    return s[:1].upper() + s[1:] if s else s


def parse_malware_from_msg(msg: str) -> Tuple[str, str]:
    if not msg:
        return "", ""
    text = _normalize_space(msg)
    low = text.lower()

    for rgx in _RE_FAMILY_STYLE:
        m = rgx.search(text)
        if m:
            fam = re.sub(r'[^A-Za-z0-9-]+', '',
                         m.group(1).strip("._-").replace("_", "-"))
            fam_title = _canon(fam)
            name = fam_title
            start = m.end()
            window = low[start:start+80]
            for kw in _SUFFIX_KEYWORDS:
                if re.search(r'\b' + re.escape(kw) + r'\b', window):
                    name = f"{fam_title} {kw.capitalize()}"
                    break
            return fam_title, name

    for k in sorted(_CANON_FAMILIES.keys(), key=len, reverse=True):
        pattern = r'\b' + re.escape(k).replace(r'\ ', r'\s*') + r'\b'
        m = re.search(pattern, low)
        if m:
            fam_title = _CANON_FAMILIES[k]
            name = fam_title
            start = m.end()
            window = low[start:start+80]
            for kw in _SUFFIX_KEYWORDS:
                if re.search(r'\b' + re.escape(kw) + r'\b', window):
                    suffix_match = re.search(
                        r'(?i)\b' + re.escape(kw) + r'\b', text[start:start+80])
                    suffix = suffix_match.group(
                        0) if suffix_match else kw.capitalize()
                    name = f"{fam_title} {suffix}"
                    break
            return fam_title, name

    m = re.search(
        r'(?i)\bET\s+(?:MALWARE|TROJAN|C2|CNC|BOTNET|RANSOMWARE)\s+([A-Za-z][A-Za-z0-9._-]{2,})(?:\s+([A-Za-z][A-Za-z0-9._-]{2,}))?', text)
    if m:
        fam_title = _canon(m.group(1))
        suffix = m.group(2)
        if suffix and suffix.lower() not in {"possible", "generic"}:
            name = f"{fam_title} {suffix}"
        else:
            start = m.end(1)
            window = low[start:start+60]
            name = fam_title
            for kw in _SUFFIX_KEYWORDS:
                if re.search(r'\b' + re.escape(kw) + r'\b', window):
                    name = f"{fam_title} {kw.capitalize()}"
                    break
        return fam_title, name

    m = re.search(r'([A-Za-z][A-Za-z0-9._-]{2,})\s*家族', text)
    if m:
        fam_title = _canon(m.group(1))
        return fam_title, fam_title

    m = re.search(r'([A-Za-z][A-Za-z0-9._-]{2,})\s*(勒索|木马|后门|蠕虫|挖矿)', text)
    if m:
        fam_title = _canon(m.group(1))
        return fam_title, f"{fam_title} {m.group(2)}"

    return "", ""

# —— 实用工具函数 ——


def deep_search(obj: Any, target_keys: Iterable[str]) -> Dict[str, Any]:
    res: Dict[str, Any] = {}
    tset = set(target_keys)

    def _walk(x: Any):
        if isinstance(x, dict):
            for k, v in x.items():
                if k in tset and k not in res:
                    res[k] = v
                _walk(v)
        elif isinstance(x, list):
            for it in x:
                _walk(it)
    _walk(obj)
    return res


def try_parse_json(text: str) -> Optional[Any]:
    try:
        return json.loads(text)
    except Exception:
        return None


def re_pick(pattern: str, text: str) -> Optional[str]:
    m = re.search(pattern, text or "", flags=re.IGNORECASE | re.DOTALL)
    if m:
        return (m.group(1) or "").strip()
    return None


def decode_tcp_payload_ascii(hex_with_colons: str) -> str:
    if not isinstance(hex_with_colons, str):
        return ""
    hx = re.sub(r"[^0-9a-fA-F]", "", hex_with_colons)
    if len(hx) % 2 == 1:
        hx = hx[:-1]
    try:
        return bytes.fromhex(hx).decode("latin1", errors="strict")
    except Exception:
        try:
            return bytes.fromhex(hx).decode("latin1", errors="ignore")
        except Exception:
            return ""


def hex_bytes(hex_str: Union[str, bytes]) -> bytes:
    """十六进制字符串 -> 原始字节（移除非 hex 符号，奇数位去尾）"""
    if not hex_str:
        return b""
    if isinstance(hex_str, bytes):
        try:
            hex_str = hex_str.decode("ascii", errors="ignore")
        except Exception:
            hex_str = str(hex_str)
    hx = re.sub(r"[^0-9a-fA-F]", "", str(hex_str))
    if len(hx) % 2 == 1:
        hx = hx[:-1]
    try:
        return bytes.fromhex(hx)
    except Exception:
        return b""


def b64_to_bytes_strict(b64s: Any) -> bytes:
    """base64 → bytes；先严格校验，失败再宽松；失败返回空字节。"""
    if not b64s or not isinstance(b64s, str):
        return b""
    s = re.sub(r"\s+", "", b64s)
    try:
        return base64.b64decode(s, validate=True)
    except Exception:
        try:
            return base64.b64decode(s, validate=False)
        except Exception:
            return b""


def bytes_to_text_preserve(b: bytes) -> str:
    """
    原样“保留”字节到文本：以 latin1 一比一映射到 Unicode，
    不做任何清洗/替换；方便写入 Excel 字符串而不丢失字节信息。
    """
    try:
        return b.decode("latin1", errors="strict")
    except Exception:
        return b.decode("latin1", errors="ignore")


def only_ssh_version(banner: Any) -> str:
    s = banner if isinstance(banner, str) else str(banner or "")
    m = re.search(r"(SSH-\d\.\d)", s)
    return m.group(1) if m else ""


_PROTO_REGEX = {
    "SSH": re.compile(r"\bssh\b|ssh\.protocol|SSH-\d\.\d", re.IGNORECASE),
    "FTP": re.compile(r"\bftp\b|ftp\.(?:request|response)\.(?:command|arg|code)", re.IGNORECASE),
    "DNS": re.compile(r"\bdns\b|dns\.", re.IGNORECASE),
    "SMB": re.compile(r"\bsmb\b|smb\.", re.IGNORECASE),
    "NBNS": re.compile(r"\bnbns\b|\bnetbios(?:-ns)?\b|nbns\.", re.IGNORECASE),
    "HTTP": re.compile(r"\bhttp\b|http\.", re.IGNORECASE),
}


def detect_l7_by_regex(pp_text: str, pp_json: Optional[Any]) -> Tuple[str, str]:
    text_parts: List[str] = []
    if isinstance(pp_text, str):
        text_parts.append(pp_text)
    if pp_json is not None:
        try:
            text_parts.append(json.dumps(pp_json, ensure_ascii=False))
        except Exception:
            pass
    haystack = "\n".join(text_parts)
    if not haystack:
        return "", ""
    earliest: Tuple[int, str] | None = None
    for proto, rgx in _PROTO_REGEX.items():
        m = rgx.search(haystack)
        if m:
            idx = m.start()
            if earliest is None or idx < earliest[0]:
                earliest = (idx, proto)
    if earliest:
        proto = earliest[1]
        return proto, proto
    return "", ""


def extract_from_payload_printable(pp_text: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {k: "" for k in PP_KEYS}
    if not pp_text:
        out["frame.protocols"] = ""
        out["l7Protocol"], out["proofType"] = "", ""
        return out

    pp_json = try_parse_json(pp_text)
    if pp_json is not None:
        found = deep_search(pp_json, PP_KEYS)
        out.update({k: found.get(k, "") for k in PP_KEYS})
        if not out.get("User-Agent") and not out.get("http.user_agent"):
            ua = re_pick(
                r"User-Agent\s*[:\"]\s*([^\r\n\"]+)", json.dumps(pp_json))
            if ua:
                out["User-Agent"] = ua
    else:
        def cap(key: str) -> str:
            return re_pick(rf'"{re.escape(key)}"\s*:\s*"(.*?)"', pp_text) or ""
        for k in PP_KEYS:
            out[k] = cap(k)
        if not out.get("User-Agent") and not out.get("http.user_agent"):
            out["User-Agent"] = re_pick(
                r"User-Agent\s*[:\"]\s*([^\r\n\"]+)", pp_text) or ""

    l7, proof = detect_l7_by_regex(pp_text, pp_json)
    out["l7Protocol"], out["proofType"] = l7, proof
    return out


# ------------------ 核心构建：返回 (row, payload_fallback_used) ------------------

def build_row(src: Dict[str, Any]) -> Tuple[List[Any], bool]:
    lOccurTime = src.get("lOccurTime", "")
    reportingTime = src.get("reportingTime", "")
    processTime = src.get("processTime", "")
    deviceTypeOneName = src.get("deviceTypeOneName", "")
    deviceTypeTwoName = src.get("deviceTypeTwoName", "")
    deviceName = src.get("deviceName", "")
    deviceCode = src.get("deviceCode", "")
    cDevIp = src.get("cDevIp", "")
    _id = src.get("id", "")
    cSrcIp = src.get("cSrcIp", "")
    iSrcPort = src.get("iSrcPort", "")
    srcCountry = src.get("srcCountry", "")
    sourceProvince = src.get("sourceProvince", src.get("srcIpProvince", ""))
    sourceCity = src.get("sourceCity", "")
    cDstIp = src.get("cDstIp", "")
    iDstPort = src.get("iDstPort", "")
    dstCountry = src.get("dstCountry", "")
    dstIpProvince = src.get("dstIpProvince", src.get("dstProvince", ""))
    ceventType = src.get("ceventType", "")
    sid = src.get("sid", "")
    ceventName = src.get("ceventName", "")
    cEventMsg = src.get("cEventMsg", "")
    cprotocol = src.get("cprotocol", "")
    eventState = src.get("eventState", "")
    cActions = src.get("cActions", "")
    cAttack = src.get("cAttack", "")
    ceventSType = src.get("ceventSType", "")
    ceventLevel = src.get("ceventLevel", "")

    virus_family, virus_name = parse_malware_from_msg(cEventMsg)

    pp_text = src.get("payloadPrintable", "") or ""
    pp_map = extract_from_payload_printable(pp_text)

    # -- packetData：cRequestMsg 原值（十六进制）
    request_body_hex = src.get("cRequestMsg", "") or ""

    # -- payload / evidence：严格按“完全保留，不做任何清洗”
    payload_b64 = (src.get("payload") or "").strip()
    if payload_b64:
        # evidence：base64 → bytes → latin1 映射文本（原样保留字节）
        raw = b64_to_bytes_strict(payload_b64)
        evidence_value = bytes_to_text_preserve(raw)

        # payload：base64 → bytes → hex
        payload_for_sheet = raw.hex()
        payload_fallback_used = False
    else:
        # payload：cRequestMsg 原十六进制
        payload_for_sheet = request_body_hex
        # evidence：cRequestMsg 十六进制 → bytes → latin1 映射文本（原样保留字节）
        raw = hex_bytes(request_body_hex)
        evidence_value = bytes_to_text_preserve(raw)
        payload_fallback_used = True

    # 其余字段
    http_method = pp_map.get("http.request.method", "")
    http_full_uri = pp_map.get("http.request.full_uri", "")
    http_referer = pp_map.get("http.referer", "")
    user_agent = pp_map.get("User-Agent", "")
    http_content_type = pp_map.get("http.content_type", "")
    dns_qry_name = pp_map.get("dns.qry.name", "")
    dns_qry_type = pp_map.get("dns.qry.type", "")
    smb_cmd = pp_map.get("smb.cmd", "")
    smb_flags = pp_map.get("smb.flags", "")
    smb_flags2 = pp_map.get("smb.flags2", "")
    smb_wct = pp_map.get("smb.wct", "")
    smb_pid = pp_map.get("smb.pid", "")
    smb_uid = pp_map.get("smb.uid", "")
    smb_mid = pp_map.get("smb.mid", "")
    ssh_banner = pp_map.get("ssh.protocol", "")
    if isinstance(ssh_banner, list):
        ssh_banner = ssh_banner[0] if ssh_banner else ""
    ssh_client_version = str(ssh_banner)
    ssh_protocol_version = only_ssh_version(ssh_banner)
    data_text_lines = pp_map.get("data-text-lines", "")
    l7Protocol = pp_map.get("l7Protocol", "")
    proofType = pp_map.get("proofType", "")

    severity = ceventLevel
    cIsAlert = str(src.get("cIsAlert", "")).strip()
    confidence = "高可信" if cIsAlert == "1" else ""
    engine = "suricata"
    l3Protocol = "IP"
    raw_msg = src.get("signaturemsg", "") or src.get("signatureMsg", "")

    row = [
        lOccurTime, reportingTime, processTime,
        deviceTypeOneName, deviceTypeTwoName, deviceName,
        deviceCode, cDevIp, _id,
        cSrcIp, iSrcPort, srcCountry, sourceProvince, sourceCity,
        cDstIp, iDstPort, dstCountry, dstIpProvince,
        ceventType, sid, ceventName, cEventMsg,
        cprotocol, eventState, cActions, cAttack, ceventSType,
        ceventLevel, virus_family, virus_name,
        payload_for_sheet,
        http_method, http_full_uri, http_full_uri,
        http_referer, user_agent, http_content_type,
        dns_qry_name, dns_qry_type,
        smb_cmd, smb_flags, smb_flags2, smb_wct,
        smb_pid, smb_uid, smb_mid,
        ssh_protocol_version, ssh_client_version, data_text_lines,
        l7Protocol, proofType,
        severity, confidence,
        evidence_value,      # evidence（完全保留，转义在写入阶段处理）
        request_body_hex,    # packetData：cRequestMsg 原值
        "",                  # requestBody 占位
        engine, l3Protocol, raw_msg,
    ]
    return row, payload_fallback_used

# -------------- 映射写入（DataFrame → 模板，保留样式） --------------


def normalize(name: str) -> str:
    if name is None:
        return ""
    s = "".join(ch for ch in str(name).strip() if ch.isalnum())
    return s.lower()


def read_header_row(ws: Worksheet, row_idx: int) -> List[str]:
    max_col = ws.max_column
    headers: List[str] = []
    for c in range(1, max_col + 1):
        headers.append(ws.cell(row=row_idx, column=c).value)
    return headers


def build_mapping_from_df(df_columns: List[str], tpl_headers: List[str]) -> Dict[int, int]:
    src_map = {normalize(h): i + 1 for i, h in enumerate(df_columns) if h}
    mapping: Dict[int, int] = {}
    for j, h in enumerate(tpl_headers, start=1):
        key = normalize(h)
        if key and key in src_map:
            mapping[src_map[key]] = j
    return mapping


def copy_style(dst_cell, src_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def _next_backup_name(base: str) -> str:
    if not os.path.exists(base):
        return base
    root, ext = os.path.splitext(base)
    n = 1
    while True:
        cand = f"{root}({n}){ext}"
        if not os.path.exists(cand):
            return cand
        n += 1


# 仅转义非法 XML 字符（\x00-\x08, \x0B, \x0C, \x0E-\x1F）。
# 保留 \x09(\t), \x0A(\n), \x0D(\r) 以及 >= \x20 的字符。
def escape_illegal_xml_chars(s: str) -> str:
    if s is None:
        return ""
    out = []
    for ch in s:
        code = ord(ch)
        if code in (0x9, 0xA, 0xD) or code >= 0x20:
            out.append(ch)
        else:
            out.append(f"\\x{code:02x}")
    return "".join(out)


def fill_template_with_df(
    df: pd.DataFrame,
    template_path: str,
    save_path: str | None = None,
    do_backup_if_overwrite: bool = True,
    sheet_name: Optional[str] = None,
) -> Tuple[int, int, Optional[str]]:
    tpl_wb = load_workbook(template_path)
    if sheet_name and sheet_name in tpl_wb.sheetnames:
        tpl_ws = tpl_wb[sheet_name]
    else:
        tpl_ws = tpl_wb[tpl_wb.sheetnames[0]]

    tpl_headers_en = read_header_row(tpl_ws, 2)
    mapping = build_mapping_from_df(list(df.columns), tpl_headers_en)
    if not mapping:
        raise RuntimeError("未在模板第2行英文表头中找到与 DataFrame 列名对应的映射。")

    backup_path: Optional[str] = None
    if save_path is None and do_backup_if_overwrite:
        root, _ext = os.path.splitext(template_path)
        backup_path_raw = f"{root}_备份.xlsx"
        backup_path = _next_backup_name(backup_path_raw)
        shutil.copy2(template_path, backup_path)

    style_row_idx = 3
    write_count = 0
    for r_idx in range(len(df)):
        r_dst = style_row_idx + write_count
        for src_col_idx, tpl_col_idx in mapping.items():
            df_col_name = str(df.columns[src_col_idx - 1] or "")
            val_raw = df.iloc[r_idx, src_col_idx - 1]

            # evidence 列：仅转义非法 XML 字符，其它不变
            if normalize(df_col_name) == "evidence":
                val = "" if val_raw is None else escape_illegal_xml_chars(
                    str(val_raw))
            else:
                val = sanitize_excel(val_raw)

            dst_cell = tpl_ws.cell(row=r_dst, column=tpl_col_idx)
            dst_cell.value = val

            style_cell = tpl_ws.cell(row=style_row_idx, column=tpl_col_idx)
            copy_style(dst_cell, style_cell)

            # ---- 证据列强制自动换行（避免长/多行文本“看起来空白”）----
            try:
                header = str(tpl_headers_en[tpl_col_idx - 1] or "")
                if normalize(header) == "evidence":
                    dst_cell.alignment = copy(dst_cell.alignment)
                    dst_cell.alignment.wrap_text = True
            except Exception:
                pass

        write_count += 1

    target_path = save_path if save_path else template_path
    tpl_wb.save(target_path)
    return write_count, len(mapping), backup_path

# ------------------ 通用读写工具 & GUI ------------------


def read_hits(data: Any) -> List[Dict[str, Any]]:
    if isinstance(data, dict):
        hh = data.get("hits", {}).get("hits")
        if isinstance(hh, list):
            return hh
        if isinstance(data.get("hits"), list):
            return data["hits"]
    if isinstance(data, list):
        return data
    return []


def sanitize_excel(s: Any) -> Any:
    if s is None:
        return ""
    try:
        if pd.isna(s):
            return ""
    except Exception:
        pass
    s = str(s)
    if s.strip().lower() == "nan":
        return ""
    # 去除 Excel/OOXML 非法控制字符，避免保存失败
    return re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]", "", s)


# —— 新的严格 JSON 加载（替代原“任意 JSON/NDJSON”解析）——
def load_json_strict(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def run_gui():
    try:
        from PyQt5 import QtWidgets, QtCore
    except Exception:
        print("[ERROR] 需要 PyQt5：pip install PyQt5", file=sys.stderr)
        raise

    class MainWindow(QtWidgets.QWidget):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("国能态感转换与映射")
            self.setMinimumWidth(820)

            self.json_path: Optional[str] = None
            self.tpl_path: Optional[str] = None

            LEFT_COL_WIDTH = 160

            self.btn_pick_json = QtWidgets.QPushButton("读取（选择 JSON）")
            self.btn_pick_json.setFixedWidth(LEFT_COL_WIDTH)
            self.le_json = QtWidgets.QLineEdit()
            self.le_json.setReadOnly(True)

            self.btn_pick_tpl = QtWidgets.QPushButton("映射（选择模板Excel）")
            self.btn_pick_tpl.setFixedWidth(LEFT_COL_WIDTH)
            self.le_tpl = QtWidgets.QLineEdit()
            self.le_tpl.setReadOnly(True)

            self.btn_sheet_label = QtWidgets.QPushButton("工作表名（可选）")
            self.btn_sheet_label.setEnabled(False)
            self.btn_sheet_label.setFixedWidth(LEFT_COL_WIDTH)
            self.le_sheet = QtWidgets.QLineEdit()
            self.le_sheet.setPlaceholderText("不填则使用模板的首个工作表")

            self.btn_convert = QtWidgets.QPushButton("转换（备份并写回）")
            self.btn_convert.setEnabled(False)
            self.btn_export_log = QtWidgets.QPushButton("导出日志到文件")

            self.log_box = QtWidgets.QTextEdit()
            self.log_box.setReadOnly(True)
            self.log_box.setPlaceholderText("日志输出...")

            layout = QtWidgets.QVBoxLayout(self)

            row1 = QtWidgets.QHBoxLayout()
            row1.addWidget(self.btn_pick_json, 0)
            row1.addWidget(self.le_json, 1)
            layout.addLayout(row1)

            row2 = QtWidgets.QHBoxLayout()
            row2.addWidget(self.btn_pick_tpl, 0)
            row2.addWidget(self.le_tpl, 1)
            layout.addLayout(row2)

            row3 = QtWidgets.QHBoxLayout()
            row3.addWidget(self.btn_sheet_label, 0)
            row3.addWidget(self.le_sheet, 1)
            layout.addLayout(row3)

            row4 = QtWidgets.QHBoxLayout()
            self.btn_convert.setSizePolicy(
                QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
            self.btn_export_log.setSizePolicy(
                QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
            row4.addWidget(self.btn_convert, 1)
            row4.addWidget(self.btn_export_log, 1)
            layout.addLayout(row4)

            layout.addWidget(self.log_box, 1)

            self.btn_pick_json.clicked.connect(self.pick_json)
            self.btn_pick_tpl.clicked.connect(self.pick_tpl)
            self.btn_convert.clicked.connect(self.do_convert)
            self.btn_export_log.clicked.connect(self.export_log)

        def log(self, msg: str):
            self.log_box.append(msg)
            print(msg)

        def pick_json(self):
            path, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, "选择 JSON 文件", "", "JSON (*.json);;所有文件 (*)"
            )
            if path:
                self.json_path = path
                self.le_json.setText(path)
                self.log(f"[INFO] 已选择 JSON：{path}")
                self.update_convert_enabled()

        def pick_tpl(self):
            path, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, "选择模板 Excel（第二行英文表头）", "", "Excel 文件 (*.xlsx);;所有文件 (*)"
            )
            if path:
                self.tpl_path = path
                self.le_tpl.setText(path)
                self.log(f"[INFO] 已选择模板：{path}")
                self.update_convert_enabled()

        def update_convert_enabled(self):
            ok = bool(self.json_path) and bool(self.tpl_path)
            self.btn_convert.setEnabled(ok)
            if ok:
                self.log("[INFO] 条件满足，可进行转换。")

        def do_convert(self):
            from PyQt5 import QtCore, QtWidgets
            if not (self.json_path and self.tpl_path):
                QtWidgets.QMessageBox.warning(self, "提示", "请先选择 JSON 与 模板文件。")
                return
            try:
                self.setEnabled(False)
                QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)
                self.log("[RUN] 开始解析 JSON 并执行映射写入...")

                # 使用严格 JSON 解析
                data = load_json_strict(self.json_path)
                hits = read_hits(data)
                self.log(f"[INFO] 读取到 hits 数量：{len(hits)}")

                rows: List[List[Any]] = []
                fallback_count = 0

                for h in hits:
                    src = h.get("_source", {}) if isinstance(h, dict) else {}
                    if not isinstance(src, dict) or not src:
                        src = h if isinstance(h, dict) else {}
                    row, used_fallback = build_row(src)
                    rows.append(row)
                    if used_fallback:
                        fallback_count += 1

                df = pd.DataFrame(rows, columns=HEADERS)

                rows_written, cols_mapped, backup_path = fill_template_with_df(
                    df=df,
                    template_path=self.tpl_path,
                    save_path=None,
                    do_backup_if_overwrite=True,
                    sheet_name=(self.le_sheet.text().strip() or None)
                )

                if backup_path:
                    self.log(f"[OK] 覆盖模板前已备份：{backup_path}")
                self.log(
                    f"[OK] 已写入：{rows_written} 行 / {cols_mapped} 列 → {self.tpl_path}")

                total = len(hits)
                pct = (fallback_count / total * 100.0) if total else 0.0
                self.log(
                    f"[STATS] payload 为空而回退到 cRequestMsg 的行数：{fallback_count} / {total} ({pct:.2f}%)")

                QtWidgets.QMessageBox.information(self, "完成", "转换与映射完成。")

            except Exception as e:
                self.log(f"[ERROR] 转换失败：{e}")
                from PyQt5 import QtWidgets
                QtWidgets.QMessageBox.critical(self, "错误", f"转换失败：\n{e}")
            finally:
                from PyQt5 import QtWidgets
                QtWidgets.QApplication.restoreOverrideCursor()
                self.setEnabled(True)

        def export_log(self):
            path, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, "保存日志到文件", "转换日志.txt", "文本文件 (*.txt);;所有文件 (*)"
            )
            if not path:
                return
            try:
                with open(path, "w", encoding="utf-8") as f:
                    f.write(self.log_box.toPlainText())
                self.log(f"[OK] 日志已导出：{path}")
            except Exception as e:
                self.log(f"[ERROR] 日志导出失败：{e}")

    from PyQt5 import QtWidgets
    app = QtWidgets.QApplication(sys.argv)
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())


# 直接启动 GUI（移除 CLI 入口）
if __name__ == "__main__":
    run_gui()
